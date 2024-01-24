# -*- coding: utf-8 -*-
import os
import re
import sys
import requests
from openpyxl import Workbook
from config.conf import Global
from utils.logger import logger
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Union, Optional, Tuple


class ApiCoverage:
    def __init__(self, swagger_url: str) -> None:
        """
        Initialize the class.

        Args:
            swagger_url (str): The URL of the Swagger file.

        Returns:
            None
        """
        self._swagger_url = swagger_url
        self._utils_dir = os.path.dirname(__file__)
        self._log_dir = os.path.abspath(os.path.join(self._utils_dir, "../log"))
        self._request_log_path = os.path.abspath(os.path.join(self._log_dir, "request.log"))

    def _merge_request_logs(self) -> None:
        """
        Merge request logs into a single file.

        Returns:
            None
        """
        log_files = [file for file in os.listdir(self._log_dir) if file.startswith("request_")]
        with open(self._request_log_path, "w", encoding="utf-8") as output_file:
            for log_file in log_files:
                file_path = os.path.abspath(os.path.join(self._log_dir, log_file))
                with open(file_path, "r") as input_file:
                    output_file.write(input_file.read().strip())

                output_file.write("\n")

    def _get_requests_from_logs(self) -> List[Dict[str, Union[str, bool]]]:
        """
        Get requests from merged request logs.

        Returns:
            List[Dict[str, Union[str, bool]]]: A list of request URLs and methods.
        """
        self._merge_request_logs()

        with open(self._request_log_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        request_list = []
        for line in lines:
            if line.startswith("http"):
                items = line.split(" ")
                path = items[2]
                path = path if path.find("?") == -1 else path[: path.find("?")]
                path = path if path[-1] != "/" else path[:-1]
                request_url = {
                    "url": items[0][: items[0].rfind(":")] + path,
                    "method": items[1][1:],
                    "matched": False,
                }
                request_list.append(request_url)

        return request_list

    def _get_requests_from_swagger(self) -> Optional[Dict[str, List[Dict[str, Union[str, int]]]]]:
        """
        Get requests from Swagger documentation.

        Returns:
            Optional[Dict[str, List[Dict[str, Union[str, int]]]]]: A dictionary containing static_url_list and
            dynamic_url_list, each of which is a list of request URLs, methods, and counts.
        """
        r = requests.get(self._swagger_url, headers=Global.constants.HEADERS)

        if r.status_code == 200:
            try:
                r.json()
            except ValueError:
                logger.error(f"Parse Swagger docs error: {r.text}")
            else:
                swagger_dict = dict()
                swagger_dict["static_url_list"] = []
                swagger_dict["dynamic_url_list"] = []
                for path, api_details in r.json().get("paths", dict()).items():
                    for method, detail in api_details.items():
                        tags = detail.get("tags")
                        if "{" not in path:
                            swagger_dict["static_url_list"].append(
                                {
                                    "url": Global.constants.BASE_URL + path,
                                    "method": method.upper(),
                                    "count": 0,
                                    "tag": tags[0] if isinstance(tags, list) and len(tags) else "NULL"
                                }
                            )
                        else:
                            swagger_dict["dynamic_url_list"].append(
                                {
                                    "url": Global.constants.BASE_URL + path,
                                    "method": method.upper(),
                                    "count": 0,
                                    "tag": tags[0] if isinstance(tags, list) and len(tags) else "NULL"
                                }
                            )

                return swagger_dict
        else:
            logger.error("Cannot request Swagger URL")
            sys.exit(1)

    @staticmethod
    def _is_similar_url(request_url: str, swagger_url: str) -> bool:
        """
        Check if the request URL is similar to the Swagger URL pattern.

        Args:
            request_url (str): The request URL to check.
            swagger_url (str): The Swagger URL pattern to compare against.

        Returns:
            bool: True if the request URL matches the Swagger URL pattern, False otherwise.
        """
        reg = r"{[^/]+?}"
        url_reg = re.sub(reg, r"[^/]+?", swagger_url)
        url_reg = "^" + url_reg + "$"
        return len(re.findall(url_reg, request_url)) != 0

    def _process(self) -> Tuple[List[Dict[str, Union[str, bool]]], Dict[str, List[Dict[str, Union[str, int]]]]]:
        """
        Process API coverage by comparing requests from logs with requests from Swagger documentation.

        Returns:
            Tuple[List[Dict[str, Union[str, bool]]], Dict[str, List[Dict[str, Union[str, int]]]]]:
            1. request_list: A list of request URLs, methods, and match status.
            2. swagger_dict: A dictionary containing static_url_list and dynamic_url_list, each of which is a list of
            request URLs, methods, and counts.
        """
        request_list = self._get_requests_from_logs()
        swagger_dict = self._get_requests_from_swagger()
        for request_url in request_list:
            for swagger_url in swagger_dict["static_url_list"]:
                if (
                        not request_url["matched"]
                        and request_url["url"] == swagger_url["url"]
                        and request_url["method"] == swagger_url["method"]
                ):
                    swagger_url["count"] = swagger_url["count"] + 1
                    request_url["matched"] = True

            for swagger_url in swagger_dict["dynamic_url_list"]:
                if (
                        not request_url["matched"]
                        and ApiCoverage._is_similar_url(request_url["url"], swagger_url["url"])
                        and request_url["method"] == swagger_url["method"]
                ):
                    swagger_url["count"] = swagger_url["count"] + 1
                    request_url["matched"] = True

        return request_list, swagger_dict

    @staticmethod
    def _set_column_width(worksheet: Worksheet) -> None:
        """
        Adjusts the column width in the worksheet.

        Args:
            worksheet (Worksheet): The worksheet to adjust the column width.

        Returns:
            None
        """
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    logger.error(e)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

    def get_coverage_summary(self) -> None:
        """
        Generate and save the API coverage summary report.

        Returns:
            None
        """
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        coverage_summary_sheet = workbook.create_sheet("coverage_summary")
        fully_covered_sheet = workbook.create_sheet("fully_covered")
        likely_covered_sheet = workbook.create_sheet("likely_covered")
        never_cover_sheet = workbook.create_sheet("never_cover")
        unknown_request_sheet = workbook.create_sheet("unknown_request")

        coverage_summary_sheet.append(["ratio", "percentage (%)"])

        for sheet in [fully_covered_sheet, likely_covered_sheet]:
            sheet.append(["module", "url", "method", "cases"])

        never_cover_sheet.append(["module", "url", "method"])
        unknown_request_sheet.append(["url", "method"])

        request_list, swagger_dict = self._process()

        total_num = len(swagger_dict["static_url_list"]) + len(swagger_dict["dynamic_url_list"])
        covered_num = len([request_url["matched"] for request_url in request_list if request_url["matched"]])
        coverage_summary_sheet.append([f"{covered_num} / {total_num}", f"%.2f" % (covered_num / total_num * 100)])

        for swagger_url in swagger_dict["static_url_list"]:
            if swagger_url["count"]:
                fully_covered_sheet.append([swagger_url["tag"],
                                            swagger_url["url"],
                                            swagger_url["method"],
                                            swagger_url["count"]])
            else:
                never_cover_sheet.append([swagger_url["tag"],
                                          swagger_url["url"],
                                          swagger_url["method"]])

        for swagger_url in swagger_dict["dynamic_url_list"]:
            if swagger_url["count"]:
                likely_covered_sheet.append([swagger_url["tag"],
                                             swagger_url["url"],
                                             swagger_url["method"],
                                             swagger_url["count"]])
            else:
                never_cover_sheet.append([swagger_url["tag"],
                                          swagger_url["url"],
                                          swagger_url["method"]])

        for request_url in request_list:
            if not request_url["matched"]:
                unknown_request_sheet.append([request_url["url"],
                                              request_url["method"]])

        for sheet_name in workbook.sheetnames:
            ApiCoverage._set_column_width(workbook[sheet_name])

        report_dir = os.path.abspath(os.path.join(self._utils_dir, "../report"))
        xlsx_path = os.path.abspath(os.path.join(report_dir, "api_coverage.xlsx"))
        workbook.save(xlsx_path)
        logger.info(f"summary path: {xlsx_path}")


if __name__ == "__main__":
    # swagger_url is the link to the swagger api-docs
    ApiCoverage(swagger_url="").get_coverage_summary()
