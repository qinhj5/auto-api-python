# -*- coding: utf-8 -*-
import os
import re
import time
import pytest
import inspect
import logging
import filelock
import traceback
from utils.tunnel_shell import TunnelShell
from utils.driver_shell import DriverShell
from openpyxl import Workbook, load_workbook
from utils.mysql_connection import MysqlConnection
from utils.redis_connection import RedisConnection
from utils.clickhouse_connection import ClickhouseConnection
from utils.dirs import log_request_dir, log_summary_dir, lock_dir, report_sheet_dir
from utils.common import (
    dumps_json,
    get_code_modifiers,
    set_column_max_width,
    get_current_datetime,
    set_allure_and_console_output,
)

session_start_time = time.time()
conftest_lock = filelock.FileLock(os.path.abspath(os.path.join(lock_dir, f"conftest.lock")))


@pytest.fixture(scope="session")
def tunnel():
    tunnel = TunnelShell()
    yield tunnel
    tunnel.close()


@pytest.fixture(scope="session")
def driver():
    driver = DriverShell()
    yield driver
    driver.close()


@pytest.fixture(scope="session")
def db():
    db = MysqlConnection()
    yield db
    db.close()


@pytest.fixture(scope="session")
def sr():
    sr = RedisConnection()
    yield sr
    sr.close()


@pytest.fixture(scope="session")
def ck():
    ck = ClickhouseConnection()
    yield ck
    ck.close()


@pytest.fixture(scope="function", autouse=True)
def case_info(request):
    set_allure_and_console_output(name="start time", body=get_current_datetime())

    func = request.function
    file_path = func.__code__.co_filename
    func_name = func.__name__
    source_code, start_line = inspect.getsourcelines(func)
    line_range = {"start_line": start_line, "end_line": start_line + len(source_code) - 1}

    set_allure_and_console_output(name="file path", body=f"{file_path}")
    set_allure_and_console_output(name="function name", body=f"{func_name}")
    set_allure_and_console_output(name="last modified by", body=get_code_modifiers(file_path, line_range))

    yield

    set_allure_and_console_output(name="end time", body=get_current_datetime())


def pytest_runtest_makereport(item, call):
    if call.when != "call" or call.excinfo is None:
        return

    reruns = item.parent.config.getoption("--reruns", 0)
    has_execution_count = hasattr(item, "execution_count")

    if not has_execution_count or (has_execution_count and item.execution_count != (reruns + 1)):
        return

    file_path = item.fspath.strpath
    case_name = item.location[-1]

    error_list = traceback.format_list(traceback.extract_tb(call.excinfo.tb))
    error_list = [error.strip() for error in error_list]

    error_idx = 0
    line_number = 1
    for idx, error in enumerate(error_list):
        if file_path in error:
            line_number = re.search(r"line (\d+)", error_list[error_idx]).group(1)
            error_idx = idx
            break

    traceback_error = ("\n".join(error_list[error_idx:])).strip()
    code_modifiers = dumps_json(get_code_modifiers(file_path=file_path, line_number=line_number))

    with conftest_lock:
        xlsx_path = os.path.abspath(os.path.join(report_sheet_dir, "failed_cases.xlsx"))
        if not os.path.exists(xlsx_path):
            os.makedirs(report_sheet_dir, exist_ok=True)

            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            failure_summary_sheet = workbook.create_sheet("failure_summary")
            failure_summary_sheet.append(["file path", "case name", "code modifiers", "traceback error"])
        else:
            workbook = load_workbook(xlsx_path)

            failure_summary_sheet = workbook["failure_summary"]

        failure_summary_sheet.append([file_path, case_name, code_modifiers, traceback_error])
        set_column_max_width(failure_summary_sheet)
        workbook.save(xlsx_path)


def pytest_sessionstart():
    global session_start_time
    session_start_time = time.time()


def pytest_terminal_summary(terminalreporter, config):
    if config.pluginmanager.get_plugin("xdist"):
        if hasattr(config, "workerinput"):
            process_name = config.workerinput["workerid"]
        elif hasattr(config, "slaveinput"):
            process_name = config.slaveinput["slaveid"]
        else:
            process_name = "main"
    else:
        process_name = "main"

    if process_name == "main":
        log_name = "summary.log"
    else:
        log_name = f"summary_{process_name}.log"
    log_path = os.path.abspath(os.path.join(log_summary_dir, log_name))

    num_passed = len(terminalreporter.stats.get("passed", []))
    num_failed = len(terminalreporter.stats.get("failed", []))
    num_error = len(terminalreporter.stats.get("error", []))
    num_skipped = len(terminalreporter.stats.get("skipped", []))
    num_collected = num_passed + num_failed + num_error + num_skipped

    with conftest_lock:
        with open(log_path, "w", encoding="utf-8") as f:
            f.writelines("Total cases: {}\n".format(num_collected))
            f.writelines("Passed cases: {}\n".format(num_passed))
            f.writelines("Failed cases: {}\n".format(num_failed))
            f.writelines("Error cases: {}\n".format(num_error))
            f.writelines("Skipped cases: {}\n".format(num_skipped))

            duration = time.time() - session_start_time
            hours = int(duration // 3600)
            minutes = int((duration % 3600) // 60)
            seconds = int(duration % 60)

            formatted_duration = "{:02d}:{:02d}:{:02d}".format(hours, minutes, seconds)

            f.writelines("Elapsed time: {}".format(formatted_duration))


@pytest.fixture(scope="session", autouse=True)
def configure_logging(request):
    if request.config.pluginmanager.get_plugin("xdist"):
        if hasattr(request.config, "workerinput"):
            process_name = request.config.workerinput["workerid"]
        elif hasattr(request.config, "slaveinput"):
            process_name = request.config.slaveinput["slaveid"]
        else:
            process_name = "main"
    else:
        process_name = "main"

    if process_name == "main":
        log_name = "request.log"
    else:
        log_name = f"request_{process_name}.log"
    log_path = os.path.abspath(os.path.join(log_request_dir, log_name))

    request_file_handler = logging.FileHandler(log_path, "w", encoding="utf-8")
    request_file_handler.setLevel(logging.DEBUG)

    request_logger = logging.getLogger("urllib3")
    request_logger.setLevel(logging.DEBUG)
    request_logger.propagate = False
    request_logger.addHandler(request_file_handler)
